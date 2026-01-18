"""
Excel Preview Service

Handles conversion of Excel files to HTML for preview display.
Supports: .xlsx, .xls, .ods
"""

from pathlib import Path
from typing import Optional
import hashlib
import json

from app.core.config import get_settings
from app.core.logging import get_logger

# Optional imports - gracefully handle missing dependencies
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    xlrd = None


logger = get_logger(__name__)
settings = get_settings()


class ExcelPreviewService:
    """
    Service for converting Excel files to HTML for preview.
    """

    CACHE_VERSION = "3"
    LAZY_LOAD_BATCH_SIZE = 200
    LAZY_LOAD_SCROLL_THRESHOLD = 120

    # Supported Excel formats
    SUPPORTED_EXCEL_FORMATS = {
        '.xlsx',  # Native support via openpyxl
        '.xlsm',  # Excel Macro-Enabled Workbook
        '.xls',   # Legacy Excel format (requires xlrd)
        '.ods',   # OpenDocument Spreadsheet (openpyxl can read)
    }

    def __init__(self):
        """Initialize the Excel preview service."""
        self.settings = settings
        self.cache_dir = Path(settings.storage.cache_dir) / "excel_previews"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def convert_excel_to_html(
        self,
        file_path: str,
        force_regenerate: bool = False,
        max_rows: int = 1000,
        max_cols: int = 50
    ) -> str:
        """
        Convert an Excel file to HTML for preview.

        Args:
            file_path: Path to the Excel file
            force_regenerate: Force regeneration even if cached
            max_rows: Maximum number of rows to render per sheet
            max_cols: Maximum number of columns to render per sheet

        Returns:
            str: Path to the generated HTML file

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file is not a supported format or dependencies not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError(
                "Excel preview feature is not available. "
                "Please install openpyxl: pip install openpyxl"
            )

        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if not path.is_file():
            raise ValueError(f"Path is not a file: {file_path}")

        # Check if Excel format is supported
        if path.suffix.lower() not in self.SUPPORTED_EXCEL_FORMATS:
            raise ValueError(
                f"Unsupported Excel format: {path.suffix}. "
                f"Supported formats: {', '.join(self.SUPPORTED_EXCEL_FORMATS)}"
            )

        # Generate cache key
        cache_key = self._generate_cache_key(file_path)
        html_path = self.cache_dir / f"{cache_key}.html"

        # Return cached HTML if exists and not forcing regeneration
        if html_path.exists() and not force_regenerate:
            logger.debug(
                f"Using cached Excel preview: {html_path}",
                extra={"source": file_path, "cached": True}
            )
            return str(html_path)

        # Convert Excel to HTML
        try:
            # Handle .xls files with xlrd if available
            if path.suffix.lower() == '.xls':
                if not XLRD_AVAILABLE:
                    raise ValueError(
                        "Cannot preview .xls files. Please install xlrd: pip install xlrd"
                    )
                html_content = self._xls_to_html(path, max_rows, max_cols)
            else:
                # .xlsx, .xlsm, .ods files - try pandas first for robustness
                logger.info(f"Loading workbook: {path}")

                try:
                    # Try pandas first - it's more robust with problematic Excel files
                    import pandas as pd
                    logger.info(f"Attempting to load with pandas...")

                    # Read all sheets
                    excel_file = pd.ExcelFile(path, engine='openpyxl')
                    html_content = self._pandas_to_html(excel_file, path.name, max_rows, max_cols)
                    excel_file.close()
                    logger.info(f"Successfully loaded with pandas")

                except Exception as pandas_error:
                    logger.warning(f"Pandas failed, falling back to openpyxl: {pandas_error}")

                    # Fallback to openpyxl
                    try:
                        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
                        logger.info(f"Workbook loaded in read-only mode")
                    except Exception as e:
                        logger.warning(f"Read-only mode failed, trying normal mode: {e}")
                        workbook = openpyxl.load_workbook(path, data_only=True, keep_links=False)
                        logger.info(f"Workbook loaded in normal mode")

                    html_content = self._workbook_to_html(workbook, path.name, max_rows, max_cols)
                    workbook.close()
                    logger.info(f"Workbook closed successfully")

            # Save HTML file
            html_path.write_text(html_content, encoding='utf-8')

            logger.info(
                f"Generated Excel preview: {html_path}",
                extra={"source": file_path}
            )

            return str(html_path)

        except Exception as e:
            logger.error(
                f"Failed to convert Excel to HTML: {file_path}: {e}",
                exc_info=True
            )
            raise ValueError(f"Failed to convert Excel file: {e}")

    def _xls_to_html(self, path: Path, max_rows: int, max_cols: int) -> str:
        """
        Convert legacy .xls file to HTML using xlrd.

        Args:
            path: Path to the .xls file
            max_rows: Maximum rows to render
            max_cols: Maximum columns to render

        Returns:
            str: HTML content
        """
        workbook = xlrd.open_workbook(path)

        html_parts = self._html_header(path.name)

        # Process each sheet
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)

            if sheet_idx > 0:
                html_parts.append('<div class="sheet-separator"></div>')

            html_parts.append(f'<div class="sheet-header">{self._escape_html(sheet.name)}</div>')

            actual_rows = min(sheet.nrows, max_rows)
            actual_cols = min(sheet.ncols, max_cols)

            if sheet.nrows == 0 or sheet.ncols == 0:
                html_parts.append('<p class="empty-sheet">Empty sheet</p>')
                continue

            header_values = []
            for col_idx in range(actual_cols):
                try:
                    cell = sheet.cell(0, col_idx)
                    header_values.append(cell.value)
                except Exception:
                    header_values.append('')

            data_rows = []
            for row_idx in range(1, actual_rows):
                row_values = []
                for col_idx in range(actual_cols):
                    try:
                        cell = sheet.cell(row_idx, col_idx)
                        row_values.append(cell.value)
                    except Exception:
                        row_values.append('')
                data_rows.append(row_values)

            html_parts.extend(self._build_lazy_table(sheet_idx, header_values, data_rows, actual_cols))

            # Show truncation notice
            if sheet.nrows > max_rows or sheet.ncols > max_cols:
                html_parts.append(
                    f'<p class="truncation-notice">Showing {actual_rows} of {sheet.nrows} rows, '
                    f'{actual_cols} of {sheet.ncols} columns</p>'
                )

        html_parts.extend(self._html_footer())
        return '\n'.join(html_parts)

    def _pandas_to_html(
        self,
        excel_file,
        filename: str,
        max_rows: int,
        max_cols: int
    ) -> str:
        """
        Convert pandas ExcelFile to HTML.

        Args:
            excel_file: pandas ExcelFile object
            filename: Original filename for display
            max_rows: Maximum rows to render
            max_cols: Maximum columns to render

        Returns:
            str: HTML content
        """
        import pandas as pd

        html_parts = self._html_header(filename)

        # Process each sheet
        for sheet_idx, sheet_name in enumerate(excel_file.sheet_names):
            if sheet_idx > 0:
                html_parts.append('<div class="sheet-separator"></div>')

            html_parts.append(f'<div class="sheet-header">{self._escape_html(sheet_name)}</div>')

            try:
                # Read sheet with pandas
                df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=max_rows)

                if df.empty:
                    html_parts.append('<p class="empty-sheet">Empty sheet</p>')
                    continue

                # Limit columns
                if len(df.columns) > max_cols:
                    df = df.iloc[:, :max_cols]
                    truncated_cols = True
                else:
                    truncated_cols = False

                header_values = list(df.columns)
                data_rows = [list(row) for row in df.itertuples(index=False, name=None)]

                html_parts.extend(
                    self._build_lazy_table(sheet_idx, header_values, data_rows, len(df.columns))
                )

                # Truncation notice
                if len(df) >= max_rows or truncated_cols:
                    total_rows = len(df)
                    total_cols = len(df.columns) if not truncated_cols else max_cols
                    html_parts.append(
                        f'<p class="truncation-notice">Showing {len(df)} rows, '
                        f'{total_cols} columns (may be truncated)</p>'
                    )

            except Exception as e:
                logger.warning(f"Failed to read sheet {sheet_name}: {e}")
                html_parts.append(f'<p class="empty-sheet">Error reading sheet: {self._escape_html(str(e))}</p>')

        html_parts.extend(self._html_footer())
        return '\n'.join(html_parts)

    def _workbook_to_html(
        self,
        workbook,
        filename: str,
        max_rows: int,
        max_cols: int
    ) -> str:
        """
        Convert an openpyxl workbook to HTML.

        Args:
            workbook: openpyxl Workbook object
            filename: Original filename for display
            max_rows: Maximum rows to render
            max_cols: Maximum columns to render

        Returns:
            str: HTML content
        """
        html_parts = self._html_header(filename)

        # Process each sheet
        for sheet_idx, sheet in enumerate(workbook.worksheets):
            if sheet_idx > 0:
                html_parts.append('<div class="sheet-separator"></div>')

            html_parts.append(f'<div class="sheet-header">{self._escape_html(sheet.title)}</div>')

            # Get actual dimensions
            if sheet.max_row == 0 or sheet.max_column == 0:
                html_parts.append('<p class="empty-sheet">Empty sheet</p>')
                continue

            actual_rows = min(sheet.max_row, max_rows)
            actual_cols = min(sheet.max_column, max_cols)

            header_values = []
            for col_idx in range(1, actual_cols + 1):
                try:
                    cell = sheet.cell(1, col_idx)
                    header_values.append(cell.value)
                except Exception:
                    header_values.append('')

            data_rows = []
            for row_idx in range(2, actual_rows + 1):
                row_values = []
                for col_idx in range(1, actual_cols + 1):
                    try:
                        cell = sheet.cell(row_idx, col_idx)
                        row_values.append(cell.value)
                    except Exception:
                        row_values.append('')
                data_rows.append(row_values)

            html_parts.extend(self._build_lazy_table(sheet_idx, header_values, data_rows, actual_cols))

            # Show truncation notice
            if sheet.max_row > max_rows or sheet.max_column > max_cols:
                html_parts.append(
                    f'<p class="truncation-notice">Showing {actual_rows} of {sheet.max_row} rows, '
                    f'{actual_cols} of {sheet.max_column} columns</p>'
                )

        html_parts.extend(self._html_footer())
        return '\n'.join(html_parts)

    def _html_header(self, filename: str) -> list:
        """Generate HTML header with styles."""
        font_stack = ', '.join([
            '"Segoe UI"',
            '"Noto Sans"',
            '"Noto Sans CJK SC"',
            '"Noto Sans CJK TC"',
            '"Noto Sans CJK JP"',
            '"Noto Sans CJK KR"',
            '"Microsoft YaHei UI"',
            '"Microsoft JhengHei UI"',
            '"Malgun Gothic"',
            '"Yu Gothic UI"',
            '"Meiryo"',
            '"Nirmala UI"',
            '"Leelawadee UI"',
            '"Arial Unicode MS"',
            '"Segoe UI Emoji"',
            '"Segoe UI Symbol"',
            '"Apple Color Emoji"',
            '"Noto Color Emoji"',
            'Arial',
            'sans-serif',
        ])
        return [
            '<!DOCTYPE html>',
            '<html>',
            '<head>',
            '<meta charset="utf-8">',
            f'<title>{self._escape_html(filename)}</title>',
            '<style>',
            f'body {{ font-family: {font_stack}; margin: 0; padding: 8px 20px 40px 20px; background: #f9fafb; user-select: text; -webkit-user-select: text; overflow-x: hidden; }}',
            '.sheet-header { font-size: 1.2em; font-weight: 600; margin: 8px 0 8px 0; color: #1e40af; padding: 10px; background: white; border-radius: 6px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }',
            '.sheet-separator { height: 20px; }',
            '.table-wrapper { overflow-x: auto; overflow-y: auto; max-height: 80vh; background: white; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); padding: 20px; margin-bottom: 40px; border: 1px solid #e2e8f0; -webkit-overflow-scrolling: touch; }',
            '.table-wrapper::-webkit-scrollbar { width: 10px; height: 10px; }',
            '.table-wrapper::-webkit-scrollbar-track { background: #f1f5f9; }',
            '.table-wrapper::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 5px; }',
            '.table-wrapper::-webkit-scrollbar-thumb:hover { background: #94a3b8; }',
            '.table-wrapper::-webkit-scrollbar-corner { background: #f1f5f9; }',
            'table { border-collapse: collapse; width: auto; min-width: 100%; font-size: 0.9em; user-select: text; table-layout: auto; }',
            'td { border: 1px solid #cbd5e1; padding: 8px 12px; white-space: nowrap; user-select: text; cursor: text; min-width: 80px; text-align: start; }',
            '.header-cell { background-color: #e0e7ff; font-weight: 600; color: #1e40af; }',
            '.data-cell { background-color: white; }',
            'tr:hover .data-cell { background-color: #f8fafc; }',
            '.empty-sheet { color: #64748b; font-style: italic; padding: 20px; }',
            '.truncation-notice { color: #64748b; font-size: 0.85em; margin-top: 10px; font-style: italic; }',
            '</style>',
            '</head>',
            '<body>',
        ]

    def _html_footer(self) -> list:
        """Generate HTML footer."""
        return [
            '<script>',
            '(function() {',
            f'const CHUNK_SIZE = {self.LAZY_LOAD_BATCH_SIZE};',
            f'const SCROLL_THRESHOLD = {self.LAZY_LOAD_SCROLL_THRESHOLD};',
            'function parseData(element) {',
            '  if (!element) {',
            '    return [];',
            '  }',
            '  try {',
            '    return JSON.parse(element.textContent || "[]");',
            '  } catch (e) {',
            '    return [];',
            '  }',
            '}',
            'function createCell(value) {',
            '  const td = document.createElement("td");',
            '  td.className = "data-cell";',
            '  td.setAttribute("dir", "auto");',
            '  td.textContent = value == null ? "" : String(value);',
            '  return td;',
            '}',
            'function appendRows(state) {',
            '  if (state.rendering || state.rendered >= state.total) {',
            '    return;',
            '  }',
            '  state.rendering = true;',
            '  requestAnimationFrame(function() {',
            '    const frag = document.createDocumentFragment();',
            '    const end = Math.min(state.rendered + CHUNK_SIZE, state.total);',
            '    for (let i = state.rendered; i < end; i++) {',
            '      const row = state.rows[i] || [];',
            '      const tr = document.createElement("tr");',
            '      for (let c = 0; c < state.colCount; c++) {',
            '        tr.appendChild(createCell(row[c]));',
            '      }',
            '      frag.appendChild(tr);',
            '    }',
            '    state.tbody.appendChild(frag);',
            '    state.rendered = end;',
            '    state.rendering = false;',
            '    if (state.rendered < state.total && state.wrapper.scrollHeight <= state.wrapper.clientHeight + SCROLL_THRESHOLD) {',
            '      appendRows(state);',
            '    }',
            '  });',
            '}',
            'function onScroll(state) {',
            '  if (state.rendered >= state.total) {',
            '    return;',
            '  }',
            '  if (state.wrapper.scrollTop + state.wrapper.clientHeight >= state.wrapper.scrollHeight - SCROLL_THRESHOLD) {',
            '    appendRows(state);',
            '  }',
            '}',
            'function initSheet(sheetId) {',
            '  const table = document.getElementById("sheet-table-" + sheetId);',
            '  const tbody = document.getElementById("sheet-body-" + sheetId);',
            '  const dataScript = document.getElementById("sheet-data-" + sheetId);',
            '  if (!table || !tbody || !dataScript) {',
            '    return;',
            '  }',
            '  const wrapper = table.closest(".table-wrapper");',
            '  if (!wrapper) {',
            '    return;',
            '  }',
            '  const colCount = parseInt(table.getAttribute("data-col-count") || "0", 10);',
            '  const rows = parseData(dataScript);',
            '  const state = {',
            '    rows: rows,',
            '    total: rows.length,',
            '    rendered: 0,',
            '    rendering: false,',
            '    colCount: colCount,',
            '    tbody: tbody,',
            '    wrapper: wrapper',
            '  };',
            '  dataScript.remove();',
            '  appendRows(state);',
            '  wrapper.addEventListener("scroll", function() {',
            '    onScroll(state);',
            '  });',
            '}',
            'function initAll() {',
            "  const scripts = document.querySelectorAll(\"script[type='application/json'][id^='sheet-data-']\");",
            '  for (let i = 0; i < scripts.length; i++) {',
            '    const id = scripts[i].id.replace("sheet-data-", "");',
            '    initSheet(id);',
            '  }',
            '}',
            'if (document.readyState === "loading") {',
            '  document.addEventListener("DOMContentLoaded", initAll);',
            '} else {',
            '  initAll();',
            '}',
            '})();',
            '</script>',
            '</body>',
            '</html>'
        ]

    def _format_cell_value(self, value) -> str:
        """Format cell value for display."""
        if value is None:
            return ''
        if isinstance(value, float):
            # Check for NaN
            import math
            if math.isnan(value):
                return ''
            # Format numbers nicely
            if value.is_integer():
                return str(int(value))
            return f'{value:.2f}'
        if isinstance(value, bool):
            return str(value)
        return str(value)

    def _render_cell(self, value, cell_class: str) -> str:
        """Render a cell with safe HTML and automatic text direction."""
        cell_value = self._format_cell_value(value)
        return f'<td class="{cell_class}" dir="auto">{self._escape_html(str(cell_value))}</td>'

    def _normalize_row(self, values, target_cols: int) -> list:
        """Normalize row values to a consistent column count."""
        row = [self._format_cell_value(val) for val in values]
        if len(row) < target_cols:
            row.extend([''] * (target_cols - len(row)))
        elif len(row) > target_cols:
            row = row[:target_cols]
        return row

    def _serialize_rows_for_js(self, rows: list) -> str:
        """Serialize rows as JSON safe for embedding in HTML."""
        serialized = json.dumps(rows)
        return serialized.replace('</', '<\\/')

    def _build_lazy_table(
        self,
        sheet_index: int,
        header_values: list,
        data_rows: list,
        column_count: int
    ) -> list:
        """Build lazy-loaded table markup for a sheet."""
        header_row = self._normalize_row(header_values, column_count)
        normalized_rows = [self._normalize_row(row, column_count) for row in data_rows]
        data_json = self._serialize_rows_for_js(normalized_rows)
        table_id = f"sheet-table-{sheet_index}"

        html_parts = [
            '<div class="table-wrapper">',
            f'<table id="{table_id}" data-sheet-index="{sheet_index}" data-col-count="{column_count}">',
            '<thead>',
            '<tr>',
        ]

        for cell in header_row:
            html_parts.append(self._render_cell(cell, 'header-cell'))

        html_parts.extend([
            '</tr>',
            '</thead>',
            f'<tbody id="sheet-body-{sheet_index}"></tbody>',
            '</table>',
            '</div>',
            f'<script type="application/json" id="sheet-data-{sheet_index}">{data_json}</script>',
        ])

        return html_parts

    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters."""
        return (str(text)
                .replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('"', '&quot;')
                .replace("'", '&#39;'))

    def is_excel(self, file_path: str) -> bool:
        """
        Check if a file is a supported Excel format.

        Args:
            file_path: Path to the file

        Returns:
            bool: True if file is a supported Excel file
        """
        path = Path(file_path)
        return path.suffix.lower() in self.SUPPORTED_EXCEL_FORMATS

    def get_cached_preview(self, file_path: str) -> Optional[str]:
        """
        Get cached preview path if it exists.

        Args:
            file_path: Path to the original file

        Returns:
            Optional[str]: Path to cached preview, or None if not cached
        """
        cache_key = self._generate_cache_key(file_path)
        html_path = self.cache_dir / f"{cache_key}.html"

        if html_path.exists():
            return str(html_path)
        return None

    def clear_cache(self, file_path: Optional[str] = None) -> int:
        """
        Clear Excel preview cache.

        Args:
            file_path: Specific file to clear cache for (None clears all)

        Returns:
            int: Number of cache files deleted
        """
        if file_path:
            # Clear specific file's preview
            cache_key = self._generate_cache_key(file_path)
            html_path = self.cache_dir / f"{cache_key}.html"

            if html_path.exists():
                html_path.unlink()
                logger.info(f"Cleared Excel preview cache for {file_path}")
                return 1
            return 0
        else:
            # Clear all previews
            count = 0
            for html_file in self.cache_dir.glob("*.html"):
                html_file.unlink()
                count += 1

            logger.info(f"Cleared all Excel preview cache ({count} files)")
            return count

    def _generate_cache_key(self, file_path: str) -> str:
        """
        Generate a unique cache key for an Excel preview.

        Args:
            file_path: Source file path

        Returns:
            str: Cache key (hash)
        """
        # Include file path and modification time in key
        path = Path(file_path)
        mtime = path.stat().st_mtime if path.exists() else 0

        key_string = f"{file_path}_{mtime}_{self.CACHE_VERSION}"
        return hashlib.md5(key_string.encode()).hexdigest()


# Singleton instance
excel_preview_service = ExcelPreviewService()
