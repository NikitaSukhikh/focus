"""
Excel operations tool for creating, editing, and reading Excel spreadsheets.

Provides comprehensive Excel manipulation with proper error handling and validation.
"""

import os
from pathlib import Path
from typing import Dict, Any, Optional, List, Union
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelOperationsTool:
    """Tool for performing Excel operations like create, read, edit spreadsheets."""

    def __init__(self):
        """Initialize the Excel operations tool."""
        self.max_file_size = 50 * 1024 * 1024  # 50 MB limit for reading

    async def create_excel(
        self,
        file_path: str,
        data: List[List[Any]],
        sheet_name: str = "Sheet1",
        has_header: bool = True
    ) -> Dict[str, Any]:
        """
        Create a new Excel file with data.

        Args:
            file_path: Path where Excel file should be saved
            data: 2D list of data (rows and columns)
            sheet_name: Name for the worksheet
            has_header: Whether first row should be formatted as header

        Returns:
            Dictionary with operation result
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            path = Path(file_path).resolve()

            # Create parent directories if needed
            path.parent.mkdir(parents=True, exist_ok=True)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Add data
            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                    # Format header row
                    if row_idx == 1 and has_header:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(str(path))

            file_size = path.stat().st_size

            return {
                "success": True,
                "path": str(path),
                "sheet_name": sheet_name,
                "rows": len(data),
                "columns": len(data[0]) if data else 0,
                "size": file_size,
                "size_human": self._format_size(file_size),
                "message": f"Successfully created Excel file: {path.name}"
            }

        except ImportError as e:
            logger.error(f"Missing dependency for Excel creation: {e}")
            return {
                "success": False,
                "error": "openpyxl not installed. Run: pip install openpyxl"
            }
        except Exception as e:
            logger.error(f"Error creating Excel {file_path}: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }

    async def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        max_rows: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Read data from an Excel file.

        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to read (None = active sheet)
            max_rows: Maximum number of rows to read (None = all)

        Returns:
            Dictionary with spreadsheet data and metadata
        """
        try:
            from openpyxl import load_workbook

            path = Path(file_path).resolve()

            # Validate file exists
            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}"
                }

            # Check file size
            file_size = path.stat().st_size
            if file_size > self.max_file_size:
                return {
                    "success": False,
                    "error": f"File too large ({file_size} bytes). Max size: {self.max_file_size} bytes"
                }

            # Read workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)

            # Get sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title

            # Extract data
            data = []
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if max_rows and row_count >= max_rows:
                    break
                # Convert row tuple to list, handling None values
                data.append([cell if cell is not None else "" for cell in row])
                row_count += 1

            wb.close()

            return {
                "success": True,
                "path": str(path),
                "sheet_name": sheet_name,
                "available_sheets": wb.sheetnames,
                "data": data,
                "rows_read": row_count,
                "columns": ws.max_column,
                "total_rows": ws.max_row,
                "truncated": max_rows is not None and row_count >= max_rows,
                "size": file_size,
                "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat()
            }

        except ImportError as e:
            logger.error(f"Missing dependency for Excel reading: {e}")
            return {
                "success": False,
                "error": "openpyxl not installed. Run: pip install openpyxl"
            }
        except Exception as e:
            logger.error(f"Error reading Excel {file_path}: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }

    async def append_to_excel(
        self,
        file_path: str,
        data: List[List[Any]],
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Append rows to an existing Excel file.

        Args:
            file_path: Path to the Excel file
            data: 2D list of data to append
            sheet_name: Sheet to append to (None = active sheet)

        Returns:
            Dictionary with operation result
        """
        try:
            from openpyxl import load_workbook

            path = Path(file_path).resolve()

            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}"
                }

            # Open workbook
            wb = load_workbook(str(path))

            # Get sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found"
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Append data
            for row_data in data:
                ws.append(row_data)

            # Save workbook
            wb.save(str(path))

            file_size = path.stat().st_size

            return {
                "success": True,
                "path": str(path),
                "rows_added": len(data),
                "size": file_size,
                "message": f"Successfully appended {len(data)} rows"
            }

        except ImportError:
            return {
                "success": False,
                "error": "openpyxl not installed. Run: pip install openpyxl"
            }
        except Exception as e:
            logger.error(f"Error appending to Excel: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }

    async def update_cell(
        self,
        file_path: str,
        row: int,
        column: Union[int, str],
        value: Any,
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Update a specific cell in an Excel file.

        Args:
            file_path: Path to the Excel file
            row: Row number (1-indexed)
            column: Column number (1-indexed) or letter (e.g., 'A', 'B')
            value: New value for the cell
            sheet_name: Sheet name (None = active sheet)

        Returns:
            Dictionary with operation result
        """
        try:
            from openpyxl import load_workbook

            path = Path(file_path).resolve()

            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}"
                }

            # Open workbook
            wb = load_workbook(str(path))

            # Get sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found"
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Update cell
            if isinstance(column, str):
                cell = ws[f"{column}{row}"]
            else:
                cell = ws.cell(row=row, column=column)

            old_value = cell.value
            cell.value = value

            # Save workbook
            wb.save(str(path))

            return {
                "success": True,
                "path": str(path),
                "row": row,
                "column": column,
                "old_value": old_value,
                "new_value": value,
                "message": f"Successfully updated cell {column}{row}"
            }

        except ImportError:
            return {
                "success": False,
                "error": "openpyxl not installed. Run: pip install openpyxl"
            }
        except Exception as e:
            logger.error(f"Error updating cell in Excel: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }

    async def create_sheet(
        self,
        file_path: str,
        sheet_name: str,
        data: Optional[List[List[Any]]] = None
    ) -> Dict[str, Any]:
        """
        Create a new sheet in an existing Excel file.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name for the new sheet
            data: Optional data to populate the sheet

        Returns:
            Dictionary with operation result
        """
        try:
            from openpyxl import load_workbook

            path = Path(file_path).resolve()

            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}"
                }

            # Open workbook
            wb = load_workbook(str(path))

            # Check if sheet already exists
            if sheet_name in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' already exists"
                }

            # Create new sheet
            ws = wb.create_sheet(title=sheet_name)

            # Add data if provided
            if data:
                for row_data in data:
                    ws.append(row_data)

            # Save workbook
            wb.save(str(path))

            return {
                "success": True,
                "path": str(path),
                "sheet_name": sheet_name,
                "rows_added": len(data) if data else 0,
                "message": f"Successfully created sheet '{sheet_name}'"
            }

        except ImportError:
            return {
                "success": False,
                "error": "openpyxl not installed. Run: pip install openpyxl"
            }
        except Exception as e:
            logger.error(f"Error creating sheet in Excel: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }

    async def delete_sheet(
        self,
        file_path: str,
        sheet_name: str
    ) -> Dict[str, Any]:
        """
        Delete a sheet from an Excel file.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of sheet to delete

        Returns:
            Dictionary with operation result
        """
        try:
            from openpyxl import load_workbook

            path = Path(file_path).resolve()

            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}"
                }

            # Open workbook
            wb = load_workbook(str(path))

            # Check if sheet exists
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' not found"
                }

            # Don't allow deleting last sheet
            if len(wb.sheetnames) == 1:
                return {
                    "success": False,
                    "error": "Cannot delete the last sheet in workbook"
                }

            # Delete sheet
            del wb[sheet_name]

            # Save workbook
            wb.save(str(path))

            return {
                "success": True,
                "path": str(path),
                "sheet_name": sheet_name,
                "message": f"Successfully deleted sheet '{sheet_name}'"
            }

        except ImportError:
            return {
                "success": False,
                "error": "openpyxl not installed. Run: pip install openpyxl"
            }
        except Exception as e:
            logger.error(f"Error deleting sheet in Excel: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }

    async def get_excel_info(
        self,
        file_path: str
    ) -> Dict[str, Any]:
        """
        Get information about an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with file information
        """
        try:
            from openpyxl import load_workbook

            path = Path(file_path).resolve()

            if not path.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}"
                }

            # Read workbook (read-only for speed)
            wb = load_workbook(str(path), read_only=True, data_only=True)

            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheets_info.append({
                    "name": sheet_name,
                    "rows": ws.max_row,
                    "columns": ws.max_column
                })

            wb.close()

            file_size = path.stat().st_size

            return {
                "success": True,
                "path": str(path),
                "size": file_size,
                "size_human": self._format_size(file_size),
                "total_sheets": len(wb.sheetnames),
                "sheets": sheets_info,
                "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat()
            }

        except ImportError:
            return {
                "success": False,
                "error": "openpyxl not installed. Run: pip install openpyxl"
            }
        except Exception as e:
            logger.error(f"Error getting Excel info: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }

    def _format_size(self, size_bytes: int) -> str:
        """Format file size in human-readable format."""
        for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.2f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.2f} PB"


# Singleton instance
excel_operations_tool = ExcelOperationsTool()
